from datetime import datetime

from logging import Logger
from log import TorgiLogger

import asyncio
import threading

import ua_generator
from aiohttp_socks import ProxyConnector, ProxyType, ProxyError, ProxyConnectionError, ProxyTimeoutError

from aiohttp import ClientSession, ClientError, TCPConnector

from tenacity import retry, retry_if_exception_type, wait_random, stop_after_attempt

from bs4 import BeautifulSoup

import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

from openpyxl.workbook.workbook import Workbook
from openpyxl import load_workbook


class TorgiScraper:
	def __init__(self, logger: Logger = TorgiLogger()):
		self._url = 'https://torgi.gov.ru/new/api/public/lotcards/rss'

		self._output_file = f'excel/output/check_results_{datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}.xlsx'

		self._semaphore = asyncio.Semaphore(value=50)
		self._thread_lock = threading.Lock()

		self._logger = logger

	@property
	def _headers(self) -> dict[str: str]:
		ua = ua_generator.generate(device='desktop')
		return {
			'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
			'Accept-Language': 'ru,en;q=0.9,en-GB;q=0.8,en-US;q=0.7',
			'Connection': 'keep-alive',
			'Referer': 'https://torgi.gov.ru/new/public/lots/reg',
			'Sec-Fetch-Dest': 'document',
			'Sec-Fetch-Mode': 'navigate',
			'Sec-Fetch-Site': 'same-origin',
			'Sec-Fetch-User': '?1',
			'Upgrade-Insecure-Requests': '1',
			'User-Agent': ua.text,
			'sec-ch-ua': ua.ch.brands,
			'sec-ch-ua-mobile': ua.ch.mobile,
			'sec-ch-ua-platform': ua.ch.platform,
		}

	@property
	def _check_result_template(self) -> dict[str: str]:
		return {
			'Номер лота': '',
			'Наименование лота': '',
			'Вид торгов': '',
			'Форма проведения торгов': '',
			'Статус лота': '',
			'Электронная площадка': '',
			'Дата публикации': '',
			'Дата изменения': '',
			'Начальная цена': '',
			'Номер извещения': '',
			'Категория имущества': '',
			'VIN': '',
			'Марка': '',
			'Модель': '',
			'Год выпуска': '',
			'Государственный регистрационный знак': '',
			'Дата государственного регистрационного знака': '',
			'Пробег': '',
			'Вид транспорта': '',
			'Объем двигателя': '',
			'Мощность двигателя': '',
			'Коробка передач': '',
			'Привод': '',
			'Экологический класс': '',
			'Статус проверки': 'Успешно'
		}

	@retry(retry=retry_if_exception_type((ProxyError, ProxyConnectionError, ProxyTimeoutError, ClientError)),
	       sleep=asyncio.sleep, wait=wait_random(min=0, max=1), stop=stop_after_attempt(10), reraise=True)
	async def _make_check_request(self, vin: str) -> str:
		search_params = {
			'biddEndFrom': '',
			'biddEndTo': '',
			'pubFrom': '',
			'pubTo': '',
			'aucStartFrom': '',
			'aucStartTo': '',
			'text': vin,
			'amoOrgCode': '',
			'npa': '',
			'byFirstVersion': 'true',
		}
		# async with ProxyConnector(proxy_type=ProxyType.HTTP, host='94.103.188.163', port='13811',
		#                           username='yfy5n4', password='s4SsUv') as proxy_conn:
		proxy_conn = TCPConnector()
		async with ClientSession(connector=proxy_conn, headers=self._headers, raise_for_status=True) as session:
			async with session.get(url=self._url, params=search_params) as check_response:
				return await check_response.text()

	@staticmethod
	def _process_item_description(item_description: BeautifulSoup) -> dict[str: str]:
		item_description_dict = {}
		columns_flag = False
		for bold_element in item_description.find_all(name='b'):
			element_text = bold_element.get_text(strip=True)
			if columns_flag and 'Характеристики' not in element_text:
				key = element_text[:-1]
				item_description_dict[key] = ''

			if 'Список лотов' in element_text:
				columns_flag = True

		item_desc_strings = list(item_description.stripped_strings)
		for key in item_description_dict:
			for idx, string in enumerate(item_desc_strings):
				if key in string:
					key_value = item_desc_strings[idx + 1]
					item_description_dict[key] = key_value if key_value[:-1] not in item_description_dict else ''
					break

		characteristics_flag = False
		for string in item_desc_strings:
			if characteristics_flag:
				key, value = string.split(':', 1)
				item_description_dict[key.strip()] = value.strip()
			if 'Характеристики' in string:
				characteristics_flag = True

		for key in item_description_dict:
			if 'Дата' in key:
				if item_description_dict[key] and isinstance(item_description_dict[key], str):
					try:
						datetime_obj = datetime.strptime(item_description_dict[key], '%Y-%m-%dT%H:%M:%S.%fZ')
						item_description_dict[key] = datetime_obj.strftime('%d.%m.%Y %H:%M:%S')
					except ValueError:
						pass

		return item_description_dict

	def _process_check_response(self, vin: str, check_response: str) -> dict[str: str]:
		check_response_xml = BeautifulSoup(markup=check_response, features='lxml-xml')
		item = check_response_xml.find(name='item')
		if not item or vin not in item.get_text():
			self._logger.info(f'VIN: {vin} | Vehicle was not found at the auction')
			check_results = self._check_result_template
			check_results['VIN'] = vin
			check_results['Статус проверки'] = 'Нет данных'
			return check_results

		item_desc_element = item.find(name='description')
		if not item_desc_element:
			raise ValueError('No description for item')

		item_desc = item_desc_element.get_text()
		item_desc_html = BeautifulSoup(markup=item_desc, features='html.parser')
		self._logger.info(f'VIN: {vin} | Item description: {item_desc}')

		item_desc_dict = self._process_item_description(item_description=item_desc_html)
		check_results = self._check_result_template
		for results_dict_key in check_results:
			for item_desc_dict_key in item_desc_dict:
				if results_dict_key in item_desc_dict_key and item_desc_dict[item_desc_dict_key]:
					check_results[results_dict_key] = item_desc_dict[item_desc_dict_key]

		self._logger.info(f'VIN: {vin} | Check result: {check_results}')
		print(f'VIN: {vin} | Check result: {check_results}')
		return check_results

	def _output_check_result(self, check_result: dict[str: str]) -> None:
		with self._thread_lock:
			try:
				wb: Workbook = load_workbook(filename=self._output_file)
				ws = wb.active
				header = False
			except FileNotFoundError:
				wb = Workbook()
				ws = wb.create_sheet(title='ГИС Торги')
				header = True

			for row in dataframe_to_rows(df=pd.json_normalize(data=check_result), index=False, header=header):
				ws.append(row)

			if header:
				for sheet_name in wb.sheetnames:
					sheet = wb[sheet_name]
					if sheet.max_row == 1 and sheet.max_column == 1:
						wb.remove(sheet)

			wb.save(filename=self._output_file)

	async def check_vehicle(self, vin: str):
		async with self._semaphore:
			try:
				check_response = await self._make_check_request(vin=vin)
				check_result = await asyncio.to_thread(self._process_check_response, vin=vin, check_response=check_response)
			except Exception as e:
				check_result = self._check_result_template
				check_result['VIN'] = vin
				check_result['Статус проверки'] = 'Ошибка'
				self._logger.error(f'VIN: {vin} | Error: {type(e)} - {e}')
			finally:
				await asyncio.to_thread(self._output_check_result, check_result=check_result)


if __name__ == '__main__':
	async def main():
		torgi_scraper = TorgiScraper()
		vin_list = [
			'XTAKS045LK1178313',
			'X9W64408MJ0002729',
			'Z8PFF3A5XBA014831',
			'Z8NBAABD0L0108892',
			'Z0X219079M0767640',
			'KNAKU811DA5005300',
			'XWEGU411BL0021018',
			'Z8NBAABD0K0083816'
		]

		check_tasks = [torgi_scraper.check_vehicle(vin=vin) for vin in vin_list]
		for check_task in asyncio.as_completed(check_tasks):
			await check_task


	loop = asyncio.get_event_loop()
	loop.run_until_complete(main())
