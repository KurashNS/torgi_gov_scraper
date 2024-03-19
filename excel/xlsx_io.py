from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

import re

from openpyxl.workbook.workbook import Workbook

import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows


def get_vin_list(input_excel_file: str):
	vin_list_workbook = load_workbook(filename=input_excel_file, read_only=True)
	vin_list_sheet: Worksheet = vin_list_workbook.active

	for col in range(vin_list_sheet.min_column, vin_list_sheet.max_column + 1):
		col_name = vin_list_sheet.cell(row=1, column=col).value
		if 'VIN' in col_name:
			vin_col = col
			break
	else:
		raise ValueError('No VIN column found in input file')

	vin_list = [cell[0].value for cell in vin_list_sheet.iter_rows(min_row=2, min_col=vin_col, max_col=vin_col)
	            if re.match(pattern=r'^[^\Wioq]{17}$', string=cell[0].value)]

	return list(set(vin_list))


def output_check_result(output_file: str, check_result: dict[str: str]) -> None:
	try:
		wb: Workbook = load_workbook(filename=output_file)
		ws = wb.active
		header = False
	except FileNotFoundError:
		wb = Workbook()
		ws = wb.create_sheet(title='ГИС Торги')
		header = True

	for row in dataframe_to_rows(df=pd.json_normalize(data=check_result), index=False, header=header):
		ws.append(row)

	if header:
		for sheet_name in wb.sheetnames:
			sheet = wb[sheet_name]
			if sheet.max_row == 1 and sheet.max_column == 1:
				wb.remove(sheet)

	wb.save(filename=output_file)
